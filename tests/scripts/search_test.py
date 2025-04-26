import pandas as pd
import requests
import time
from datetime import datetime
import logging
from typing import Dict, List, Union, Optional, Tuple
import os
import sys
import json
import asyncio
import aiohttp
import platform
from concurrent.futures import ThreadPoolExecutor
import statistics
from dataclasses import dataclass
from enum import Enum
import random
from collections import defaultdict
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.label import DataLabelList

# Thiết lập event loop cho Windows
if platform.system() == 'Windows':
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

class TestMode(Enum):
    CONCURRENT = "concurrent"
    LOAD = "load"

class EndpointType(Enum):
    SEARCH = "search"  # Tìm kiếm cơ bản
    DETAIL = "detail"  # Thông tin chi tiết
    EDUCATION = "education"  # Thông tin học vấn
    EXAMS = "exams"  # Thông tin thi cử
    ACHIEVEMENTS = "achievements"  # Thông tin thành tích

class SearchCriteria(Enum):
    EXACT_MATCH = "exact_match"  # Tìm chính xác
    FUZZY_MATCH = "fuzzy_match"  # Tìm gần đúng
    PARTIAL_MATCH = "partial_match"  # Tìm một phần
    CASE_INSENSITIVE = "case_insensitive"  # Không phân biệt hoa thường
    MULTI_FIELD = "multi_field"  # Tìm nhiều trường
    INVALID = "invalid"  # Dữ liệu không hợp lệ

@dataclass
class TestConfig:
    mode: TestMode
    concurrent_requests: int = 1
    test_duration_seconds: int = 60
    batch_size: int = 50  # Số request mỗi batch
    max_connections: int = 500  # Giới hạn kết nối đồng thời

@dataclass
class TestResult:
    timestamp: datetime
    endpoint_type: str
    search_type: str
    search_params: Dict
    status_code: int
    response_time: float
    results_count: int
    error: Optional[str] = None

class SearchTest:
    def __init__(self, config: TestConfig):
        self.base_url = "http://localhost:8000/api/v1"
        self.setup_logging()
        self.results: List[TestResult] = []
        self.config = config
        self.test_data = None
        self.connector = None
        self.session = None
        self.candidate_ids = set()
        self.tested_records = set()
        self.stats = defaultdict(lambda: defaultdict(lambda: {
            'total': 0,
            'success': 0,
            'response_times': [],
            'result_counts': [],
            'errors': defaultdict(int)
        }))
        
    def setup_logging(self):
        """Thiết lập logging để ghi ra cả file và terminal"""
        log_dir = "../output/logs"
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)
            
        logger = logging.getLogger()
        logger.setLevel(logging.INFO)
        
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        
        file_handler = logging.FileHandler(
            f'{log_dir}/search_test_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log',
            encoding='utf-8'
        )
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)

        if sys.stdout.encoding != 'utf-8':
            sys.stdout.reconfigure(encoding='utf-8')

    def load_test_data(self) -> pd.DataFrame:
        """Đọc dữ liệu từ file Excel"""
        try:
            file_path = "../data/input/data_search.xlsx"
            df = pd.read_excel(file_path)
            logging.info(f"Đã đọc file {file_path} thành công")
            logging.info(f"Số lượng bản ghi: {len(df)}")
            self.test_data = df
            return df
        except Exception as e:
            logging.error(f"Lỗi khi đọc file test data: {str(e)}")
            raise

    def generate_test_cases(self, row) -> List[Dict]:
        """Tạo các test case đa dạng từ một bản ghi"""
        test_cases = []
        
        # Chuẩn bị dữ liệu gốc
        fields = {
            'candidate_id': str(row['MÃ THÍ SINH']),
            'id_number': str(row['CMND/CCCD']),
            'full_name': str(row['HỌ VÀ TÊN'])
        }
        
        # 1. Test tìm kiếm cơ bản
        basic_search_cases = self._generate_basic_search_cases(fields)
        for case in basic_search_cases:
            case['endpoint'] = EndpointType.SEARCH
        test_cases.extend(basic_search_cases)
        
        # 2. Test các endpoint thông tin chi tiết
        detail_cases = self._generate_detail_cases(fields['candidate_id'])
        test_cases.extend(detail_cases)
        
        return test_cases

    def _validate_test_data(self, fields: Dict) -> Dict:
        """Validate và chuẩn hóa dữ liệu test"""
        validated = {}
        for k, v in fields.items():
            if v is None:
                continue
            
            # Chuyển về string và chuẩn hóa
            cleaned = str(v).strip()
            if not cleaned:
                continue
                
            # Xử lý các trường đặc biệt
            if k == 'candidate_id' and not cleaned.isalnum():
                continue
            if k == 'id_number' and not any(c.isdigit() for c in cleaned):
                continue
            if k == 'full_name' and len(cleaned.split()) < 2:
                continue
                
            validated[k] = cleaned
            
        return validated

    def _generate_fuzzy_match(self, field: str, value: str) -> List[str]:
        """Tạo các biến thể tìm kiếm gần đúng thông minh hơn"""
        results = []
        
        if not value or len(value) < 3:
            return results
            
        if field == 'full_name':
            # Tách họ tên
            parts = value.split()
            if len(parts) >= 2:
                results.extend([
                    parts[0],  # Họ
                    parts[-1],  # Tên
                    ' '.join(parts[:2]),  # Họ + tên đệm
                    ' '.join(parts[-2:])  # Tên đệm + tên
                ])
        elif field in ['candidate_id', 'id_number']:
            # Xử lý ID và số
            if len(value) >= 4:
                results.extend([
                    value[:4],  # 4 ký tự đầu
                    value[-4:],  # 4 ký tự cuối
                    ''.join(c for c in value if c.isdigit())[-4:]  # 4 số cuối
                ])
        else:
            # Các trường khác
            if len(value) >= 4:
                results.extend([
                    value[:len(value)//2],  # Nửa đầu
                    value[len(value)//4:3*len(value)//4],  # Phần giữa
                    value[-4:]  # 4 ký tự cuối
                ])
        
        return list(set(results))  # Loại bỏ trùng lặp

    def _generate_basic_search_cases(self, fields: Dict) -> List[Dict]:
        """Tạo các test case thông minh hơn"""
        test_cases = []
        validated_fields = self._validate_test_data(fields)
        
        if not validated_fields:
            return test_cases
        
        # Test tìm chính xác từng trường
        for field, value in validated_fields.items():
            test_cases.append({
                'data': {field: value},
                'type': SearchCriteria.EXACT_MATCH
            })
        
        # Test tìm gần đúng
        for field, value in validated_fields.items():
            fuzzy_values = self._generate_fuzzy_match(field, value)
            for fuzzy_value in fuzzy_values:
                test_cases.append({
                    'data': {field: fuzzy_value},
                    'type': SearchCriteria.FUZZY_MATCH
                })
        
        # Test không phân biệt hoa thường
        for field, value in validated_fields.items():
            if any(c.isalpha() for c in value):
                test_cases.append({
                    'data': {field: value.upper()},
                    'type': SearchCriteria.CASE_INSENSITIVE
                })
                test_cases.append({
                    'data': {field: value.lower()},
                    'type': SearchCriteria.CASE_INSENSITIVE
                })
        
        # Test nhiều trường
        if len(validated_fields) >= 2:
            fields_list = list(validated_fields.items())
            for i in range(len(fields_list)):
                for j in range(i + 1, len(fields_list)):
                    test_cases.append({
                        'data': {
                            fields_list[i][0]: fields_list[i][1],
                            fields_list[j][0]: fields_list[j][1]
                        },
                        'type': SearchCriteria.MULTI_FIELD
                    })
        
        return test_cases

    def _generate_detail_cases(self, candidate_id: str) -> List[Dict]:
        """Tạo các test case cho các endpoint thông tin chi tiết"""
        test_cases = []
        
        # Test thông tin chi tiết cơ bản
        test_cases.append({
            'endpoint': EndpointType.DETAIL,
            'data': {'candidate_id': candidate_id},
            'type': SearchCriteria.EXACT_MATCH,
            'params': {
                'include_education': True,
                'include_exams': True,
                'include_achievements': True
            }
        })
        
        # Test thông tin học vấn
        test_cases.append({
            'endpoint': EndpointType.EDUCATION,
            'data': {'candidate_id': candidate_id},
            'type': SearchCriteria.EXACT_MATCH
        })
        
        # Test thông tin thi cử
        test_cases.append({
            'endpoint': EndpointType.EXAMS,
            'data': {'candidate_id': candidate_id},
            'type': SearchCriteria.EXACT_MATCH
        })
        
        # Test thông tin thành tích
        test_cases.append({
            'endpoint': EndpointType.ACHIEVEMENTS,
            'data': {'candidate_id': candidate_id},
            'type': SearchCriteria.EXACT_MATCH
        })
        
        return test_cases

    def categorize_test_case(self, test_case: Dict) -> SearchCriteria:
        """Phân loại test case theo tiêu chí tìm kiếm"""
        if not test_case:
            return SearchCriteria.INVALID
            
        if any(not str(v).strip() for v in test_case.values()):
            return SearchCriteria.INVALID
            
        num_fields = len(test_case)
        
        if num_fields == 1:
            return SearchCriteria.EXACT_MATCH
        elif num_fields == 2:
            return SearchCriteria.FUZZY_MATCH
        elif num_fields == 3:
            return SearchCriteria.MULTI_FIELD
        else:
            return SearchCriteria.INVALID

    async def setup_session(self):
        """Thiết lập session với các tham số phù hợp"""
        if self.connector is None:
            self.connector = aiohttp.TCPConnector(
                limit=self.config.max_connections,
                force_close=True,
                enable_cleanup_closed=True
            )
        
        if self.session is None:
            timeout = aiohttp.ClientTimeout(total=30)
            self.session = aiohttp.ClientSession(
                connector=self.connector,
                timeout=timeout
            )
        
        return self.session

    async def cleanup(self):
        """Dọn dẹp tài nguyên"""
        if self.session:
            await self.session.close()
        if self.connector:
            await self.connector.close()

    async def search_by_fields_async(self, session: aiohttp.ClientSession, test_case: Dict) -> TestResult:
        """Thực hiện request với xử lý lỗi tốt hơn"""
        start_time = time.time()
        endpoint_type = test_case.get('endpoint', EndpointType.SEARCH)
        search_type = test_case['type']
        
        try:
            # Validate params
            params = self._validate_test_data(test_case['data'])
            if not params:
                raise ValueError("No valid search parameters")
            
            # Xây dựng URL
            if endpoint_type == EndpointType.SEARCH:
                url = f"{self.base_url}/search/candidates"
            else:
                candidate_id = params.pop('candidate_id', None)
                if not candidate_id:
                    raise ValueError("Missing candidate_id for detail endpoint")
                    
                base_path = f"{self.base_url}/search/candidates/{candidate_id}"
                url = base_path
                
                if endpoint_type == EndpointType.EDUCATION:
                    url = f"{base_path}/education"
                elif endpoint_type == EndpointType.EXAMS:
                    url = f"{base_path}/exams"
                elif endpoint_type == EndpointType.ACHIEVEMENTS:
                    url = f"{base_path}/achievements"
            
            # Thực hiện request
            async with session.get(url, params=params, timeout=30) as response:
                end_time = time.time()
                response_time = end_time - start_time
                
                # Xử lý response
                data = await response.json()
                results_count = 0
                
                if response.status == 200:
                    if endpoint_type == EndpointType.SEARCH:
                        candidates = data.get('candidates', [])
                        results_count = len(candidates)
                        for candidate in candidates:
                            self.candidate_ids.add(candidate.get('candidate_id'))
                    else:
                        results_count = 1 if data else 0
                
                # Cập nhật thống kê
                stats = self.stats[endpoint_type.value][search_type.value]
                stats['total'] += 1
                if response.status == 200:
                    stats['success'] += 1
                stats['response_times'].append(response_time)
                stats['result_counts'].append(results_count)
                
                if response.status != 200:
                    error_type = data.get('error', 'UNKNOWN_ERROR')
                    stats['errors'][error_type] += 1
                
                return TestResult(
                    timestamp=datetime.now(),
                    endpoint_type=endpoint_type.value,
                    search_type=search_type.value,
                    search_params=params,
                    status_code=response.status,
                    response_time=response_time,
                    results_count=results_count,
                    error=None if response.status == 200 else json.dumps(data)
                )
                
        except Exception as e:
            end_time = time.time()
            error_msg = str(e)
            
            # Cập nhật thống kê lỗi
            stats = self.stats[endpoint_type.value][search_type.value]
            stats['total'] += 1
            stats['errors']['EXCEPTION'] += 1
            stats['response_times'].append(end_time - start_time)
            stats['result_counts'].append(0)
            
            return TestResult(
                timestamp=datetime.now(),
                endpoint_type=endpoint_type.value,
                search_type=search_type.value,
                search_params=test_case['data'],
                status_code=500,
                response_time=end_time - start_time,
                results_count=0,
                error=error_msg
            )

    def generate_excel_report(self):
        """Tạo báo cáo Excel với nhiều sheet thống kê"""
        wb = Workbook()
        
        # 1. Sheet kết quả chi tiết
        ws_detail = wb.active
        ws_detail.title = "Detailed Results"
        
        # Chuyển kết quả test thành DataFrame
        results_data = []
        for result in self.results:
            results_data.append({
                'Timestamp': result.timestamp,
                'Endpoint Type': result.endpoint_type,
                'Search Type': result.search_type,
                'Search Params': json.dumps(result.search_params),
                'Status Code': result.status_code,
                'Response Time (s)': result.response_time,
                'Results Count': result.results_count,
                'Error': result.error
            })
        
        df_results = pd.DataFrame(results_data)
        for row in dataframe_to_rows(df_results, index=False, header=True):
            ws_detail.append(row)
        
        # 2. Sheet thống kê độ chính xác
        ws_accuracy = wb.create_sheet("Accuracy Statistics")
        accuracy_data = []
        
        for endpoint in self.stats:
            for search_type in self.stats[endpoint]:
                stats = self.stats[endpoint][search_type]
                total = stats['total']
                success = stats['success']
                accuracy = (success / total * 100) if total > 0 else 0
                
                accuracy_data.append({
                    'Endpoint': endpoint,
                    'Search Type': search_type,
                    'Total Requests': total,
                    'Successful Requests': success,
                    'Success Rate (%)': accuracy,
                    'Avg Results': np.mean(stats['result_counts']) if stats['result_counts'] else 0,
                    'Max Results': max(stats['result_counts']) if stats['result_counts'] else 0,
                    'Zero Results (%)': (sum(1 for x in stats['result_counts'] if x == 0) / total * 100) if total > 0 else 0
                })
        
        df_accuracy = pd.DataFrame(accuracy_data)
        for row in dataframe_to_rows(df_accuracy, index=False, header=True):
            ws_accuracy.append(row)
        
        # Thêm biểu đồ Success Rate
        chart = BarChart()
        chart.title = "Success Rate by Search Type"
        chart.x_axis.title = "Search Type"
        chart.y_axis.title = "Success Rate (%)"
        
        data = Reference(ws_accuracy, min_col=5, min_row=2, max_row=len(accuracy_data)+1)
        cats = Reference(ws_accuracy, min_col=2, min_row=2, max_row=len(accuracy_data)+1)
        chart.add_data(data)
        chart.set_categories(cats)
        
        ws_accuracy.add_chart(chart, "J2")
        
        # 3. Sheet thống kê hiệu suất
        ws_perf = wb.create_sheet("Performance Metrics")
        perf_data = []
        
        for endpoint in self.stats:
            for search_type in self.stats[endpoint]:
                stats = self.stats[endpoint][search_type]
                response_times = stats['response_times']
                
                if response_times:
                    perf_data.append({
                        'Endpoint': endpoint,
                        'Search Type': search_type,
                        'Min Response Time (s)': min(response_times),
                        'Max Response Time (s)': max(response_times),
                        'Avg Response Time (s)': np.mean(response_times),
                        'Median Response Time (s)': np.median(response_times),
                        '90th Percentile (s)': np.percentile(response_times, 90),
                        '95th Percentile (s)': np.percentile(response_times, 95),
                        'Std Dev (s)': np.std(response_times)
                    })
        
        df_perf = pd.DataFrame(perf_data)
        for row in dataframe_to_rows(df_perf, index=False, header=True):
            ws_perf.append(row)
        
        # Thêm biểu đồ Response Time
        chart = LineChart()
        chart.title = "Response Time Statistics"
        chart.x_axis.title = "Endpoint + Search Type"
        chart.y_axis.title = "Time (s)"
        
        data = Reference(ws_perf, min_col=3, max_col=7, min_row=1, max_row=len(perf_data)+1)
        cats = Reference(ws_perf, min_col=1, max_col=2, min_row=2, max_row=len(perf_data)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws_perf.add_chart(chart, "J2")
        
        # 4. Sheet phân tích lỗi
        ws_errors = wb.create_sheet("Error Analysis")
        error_data = []
        
        for endpoint in self.stats:
            for search_type in self.stats[endpoint]:
                stats = self.stats[endpoint][search_type]
                for error_type, count in stats['errors'].items():
                    error_data.append({
                        'Endpoint': endpoint,
                        'Search Type': search_type,
                        'Error Type': error_type,
                        'Count': count,
                        'Error Rate (%)': (count / stats['total'] * 100) if stats['total'] > 0 else 0
                    })
        
        df_errors = pd.DataFrame(error_data)
        for row in dataframe_to_rows(df_errors, index=False, header=True):
            ws_errors.append(row)
        
        # Lưu file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = "../output/results"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        output_file = f"{output_dir}/search_results_{timestamp}.xlsx"
        wb.save(output_file)
        logging.info(f"\nĐã lưu báo cáo chi tiết vào: {output_file}")

    async def run_tests(self):
        """Chạy các test tìm kiếm"""
        try:
            if self.test_data is None:
                self.load_test_data()
            
            if self.config.mode == TestMode.CONCURRENT:
                await self.run_concurrent_tests(self.config.concurrent_requests)
            elif self.config.mode == TestMode.LOAD:
                await self.run_load_test(self.config.test_duration_seconds)
                
            # Kiểm tra các record chưa được test
            total_records = len(self.test_data)
            tested_records = len(self.tested_records)
            untested_records = total_records - tested_records
            
            if untested_records > 0:
                logging.warning(f"Còn {untested_records} record chưa được test!")
                # Test các record còn lại
                await self.test_remaining_records()
                
        except Exception as e:
            logging.error(f"Lỗi khi chạy test: {str(e)}")
            raise

    async def test_remaining_records(self):
        """Test các record chưa được test"""
        try:
            remaining_records = []
            for idx, row in self.test_data.iterrows():
                if idx not in self.tested_records:
                    remaining_records.append(row)
            
            if not remaining_records:
                return
                
            logging.info(f"Bắt đầu test {len(remaining_records)} record còn lại")
            
            # Tạo test cases cho các record còn lại
            all_test_cases = []
            for row in remaining_records:
                test_cases = self.generate_test_cases(row)
                all_test_cases.extend(test_cases)
                self.tested_records.add(row.name)  # Đánh dấu record đã được test
            
            # Chạy test theo batch
            batch_size = self.config.batch_size
            session = await self.setup_session()
            
            try:
                for i in range(0, len(all_test_cases), batch_size):
                    batch = all_test_cases[i:i + batch_size]
                    batch_start = time.time()
                    
                    tasks = [
                        self.search_by_fields_async(session, case)
                        for case in batch
                    ]
                    
                    results = await asyncio.gather(*tasks, return_exceptions=True)
                    
                    # Xử lý kết quả
                    batch_success = 0
                    batch_error = 0
                    valid_results = []
                    
                    for result in results:
                        if isinstance(result, Exception):
                            batch_error += 1
                            logging.error(f"Request error: {str(result)}")
                        else:
                            batch_success += 1
                            valid_results.append(result)
                    
                    self.results.extend(valid_results)
                    
                    # Thống kê batch
                    batch_time = time.time() - batch_start
                    batch_rps = len(batch) / batch_time if batch_time > 0 else 0
                    
                    logging.info(f"\nBatch cho records còn lại {i//batch_size + 1}/{(len(all_test_cases) + batch_size - 1)//batch_size}:")
                    logging.info(f"- Số lượng request: {len(batch)}")
                    logging.info(f"- Thành công: {batch_success}, Lỗi: {batch_error}")
                    logging.info(f"- Thời gian: {batch_time:.2f} giây")
                    logging.info(f"- Tốc độ: {batch_rps:.2f} requests/giây")
                    
            finally:
                await self.cleanup()
                
        except Exception as e:
            logging.error(f"Lỗi khi test các record còn lại: {str(e)}")
            raise

    async def run_concurrent_tests(self, num_requests: int):
        """Chạy test đồng thời với số lượng request chỉ định"""
        if self.test_data is None:
            self.load_test_data()
            
        all_test_cases = []
        for idx, row in self.test_data.iterrows():
            test_cases = self.generate_test_cases(row)
            all_test_cases.extend(test_cases)
            self.tested_records.add(idx)  # Đánh dấu record đã được test
            
            # Nếu đã đủ số lượng request cần thiết thì dừng
            if len(all_test_cases) >= num_requests:
                break
        
        # Nếu chưa đủ số request, lặp lại các test case
        while len(all_test_cases) < num_requests:
            all_test_cases.extend(all_test_cases[:num_requests - len(all_test_cases)])
        
        test_cases = all_test_cases[:num_requests]
        batch_size = self.config.batch_size
        
        logging.info(f"Bắt đầu chạy {num_requests} requests")
        logging.info(f"Kích thước batch: {batch_size}")
        logging.info(f"Số kết nối tối đa: {self.config.max_connections}")
        
        total_success = 0
        total_error = 0
        start_time = time.time()
        
        try:
            session = await self.setup_session()
            
            # Chia thành các batch
            for i in range(0, len(test_cases), batch_size):
                batch = test_cases[i:i + batch_size]
                batch_start = time.time()
                
                # Tạo và thực hiện các request trong batch
                tasks = [
                    self.search_by_fields_async(session, case)
                    for case in batch
                ]
                
                # Chạy các request trong batch
                results = await asyncio.gather(*tasks, return_exceptions=True)
                
                # Xử lý kết quả
                batch_success = 0
                batch_error = 0
                valid_results = []
                
                for result in results:
                    if isinstance(result, Exception):
                        batch_error += 1
                        logging.error(f"Request error: {str(result)}")
                    else:
                        batch_success += 1
                        valid_results.append(result)
                
                total_success += batch_success
                total_error += batch_error
                self.results.extend(valid_results)
                
                # Thống kê batch
                batch_time = time.time() - batch_start
                batch_rps = len(batch) / batch_time if batch_time > 0 else 0
                
                logging.info(f"\nBatch {i//batch_size + 1}/{(len(test_cases) + batch_size - 1)//batch_size}:")
                logging.info(f"- Số lượng request: {len(batch)}")
                logging.info(f"- Thành công: {batch_success}, Lỗi: {batch_error}")
                logging.info(f"- Thời gian: {batch_time:.2f} giây")
                logging.info(f"- Tốc độ: {batch_rps:.2f} requests/giây")
                
        finally:
            await self.cleanup()
        
        total_time = time.time() - start_time
        
        logging.info(f"\nKết quả tổng thể:")
        logging.info(f"Tổng số request: {num_requests}")
        logging.info(f"Thành công: {total_success}, Lỗi: {total_error}")
        logging.info(f"Thời gian: {total_time:.2f} giây")
        logging.info(f"Tốc độ trung bình: {num_requests/total_time:.2f} requests/giây")
        logging.info(f"Số record đã test: {len(self.tested_records)}/{len(self.test_data)}")

    async def run_load_test(self, duration_seconds: int = 60):
        """
        Chạy load test trong khoảng thời gian chỉ định
        Args:
            duration_seconds: Thời gian chạy test (giây)
        """
        if not self.test_data is not None:
            self.load_test_data()
            
        all_test_cases = []
        for _, row in self.test_data.iterrows():
            all_test_cases.extend(self.generate_test_cases(row))
        
        logging.info(f"Bắt đầu load test trong {duration_seconds} giây")
        start_time = time.time()
        request_count = 0
        error_count = 0
        
        connector = aiohttp.TCPConnector(limit=self.config.max_connections,
                                       force_close=True,
                                       enable_cleanup_closed=True)
        
        async with aiohttp.ClientSession(connector=connector) as session:
            while time.time() - start_time < duration_seconds:
                current_time = time.time()
                remaining_time = duration_seconds - (current_time - start_time)
                
                if remaining_time <= 0:
                    break
                
                # Tạo nhiều request cùng lúc
                batch_size = min(100, self.config.max_connections)
                tasks = []
                
                for _ in range(batch_size):
                    test_case = all_test_cases[request_count % len(all_test_cases)]
                    task = asyncio.create_task(self.search_by_fields_async(session, test_case))
                    tasks.append(task)
                    request_count += 1
                
                # Chạy các request và xử lý kết quả
                results = await asyncio.gather(*tasks, return_exceptions=True)
                
                for result in results:
                    if isinstance(result, Exception):
                        error_count += 1
                        logging.error(f"Request error: {str(result)}")
                        continue
                    self.results.append(result)
                
                # Đợi một chút để tránh quá tải CPU
                await asyncio.sleep(0.01)
        
        end_time = time.time()
        total_time = end_time - start_time
        successful_requests = request_count - error_count
        
        logging.info(f"Hoàn thành {successful_requests}/{request_count} requests trong {total_time:.2f} giây")
        logging.info(f"Tốc độ trung bình: {successful_requests/total_time:.2f} requests/giây")
        logging.info(f"Tỷ lệ lỗi: {(error_count/request_count)*100:.1f}%")

async def main():
    # Các cấu hình test với số lượng request tăng dần
    test_configs = [
        TestConfig(
            mode=TestMode.CONCURRENT,
            concurrent_requests=100,
            batch_size=20,
            max_connections=100
        ),
        TestConfig(
            mode=TestMode.CONCURRENT,
            concurrent_requests=500,
            batch_size=50,
            max_connections=200
        ),
        TestConfig(
            mode=TestMode.CONCURRENT,
            concurrent_requests=1000,
            batch_size=100,
            max_connections=300
        )
    ]
    
    for config in test_configs:
        logging.info(f"\n=== BẮT ĐẦU TEST VỚI {config.concurrent_requests} REQUESTS ===")
        test = SearchTest(config)
        await test.run_tests()
        test.generate_excel_report()
        logging.info(f"=== KẾT THÚC TEST VỚI {config.concurrent_requests} REQUESTS ===\n")
        await asyncio.sleep(2)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except Exception as e:
        logging.error(f"Lỗi nghiêm trọng: {str(e)}")
        sys.exit(1) 